import json

import openpyxl

import effect_def
from type_def import Card


def load_from_excel(cs: list[Card]):
    # dict_name_cpf = {}
    # for attr in dir(card_func_def):
    #     ele = getattr(card_func_def, attr)
    #     if callable(ele):
    #         name = ele.__name__
    #         if len(name) > 3 and name[:3] == 'cp_':
    #             dict_name_cpf[name[3:]] = ele

    book = openpyxl.load_workbook('adv_ire.xlsx')
    sheet = book['card']
    for i in range(2, 100):
        name = sheet.cell(row=i, column=1).value
        label = sheet.cell(row=i, column=4).value
        ej = sheet.cell(row=i, column=8).value
        if ej is None or len(ej) == 0:
            break
        ef = []
        for j in json.loads(ej):
            for k, v in j.items():
                ef.append(effect_def.atom_effect_name_to_compile_function[k](v))
        cs.append(Card(name, label, ef))


if __name__ == '__main__':
    cards = []
    load_from_excel(cards)
    for c in cards:
        print(c.name)
        for cc in c.effects:
            print(cc.select_unit)
            print(cc.effect_func)
            print(cc.effect_param)
        print('-' * 8)
